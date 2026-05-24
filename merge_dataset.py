import openpyxl, json, re, os

folder = r'd:\Sardesai_Data\MINOR\MINOR\CASE_SEARCH_DATASET'
out_path = r'd:\Sardesai_Data\MINOR\MINOR\backend\src\data\legal_statutes.json'

def clean(val):
    if not val: return ''
    return re.sub(r'\n.*', '', str(val)).strip()

seen = set()
data = []

def add(category, offense, ipc, bns, description=''):
    # Normalize fragmented tort categories into one group
    tort_types = ['Battery','Assault','False Imprisonment','Conversion / Trespass to Goods',
                  'Criminal Negligence','Fraud / Deceit','Intentional Insult / Nervous Shock',
                  'Malicious Prosecution','Public Nuisance','Trespass to Land']
    if category in tort_types:
        category = 'Torts / Civil Wrongs'
    key = (category.strip(), offense.strip())
    if key in seen or not offense.strip() or not category.strip():
        return
    seen.add(key)
    data.append({
        'category': category.strip(),
        'offense': offense.strip(),
        'ipc': ipc.strip() if ipc else '',
        'bns': bns.strip() if bns else '',
        'description': description.strip() if description else ''
    })

# 1. Legal_Statutes_Simple_Marathi.xlsx — cols: category, offense, ipc, bns
wb = openpyxl.load_workbook(os.path.join(folder, 'Legal_Statutes_Simple_Marathi.xlsx'))
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    if not any(row): continue
    add(clean(row[0]), clean(row[1]), clean(row[2]), clean(row[3]))

# 2. DOMESTIC.xlsx — cols: category, offense, ipc, bns
wb = openpyxl.load_workbook(os.path.join(folder, 'DOMESTIC.xlsx'))
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    if not any(row): continue
    add(clean(row[0]), clean(row[1]), clean(row[2]), clean(row[3]))

# 3. moredataVIOLENT.xlsx — cols: #, offense, description, ipc, bns
wb = openpyxl.load_workbook(os.path.join(folder, 'moredataVIOLENT.xlsx'))
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    if not any(row): continue
    add('Violent Crimes', clean(row[1]), clean(row[3]), clean(row[4]), clean(row[2]))

# 4. Property_Crime.xlsx — cols: category, offense, ipc, bns
wb = openpyxl.load_workbook(os.path.join(folder, 'Property_Crime.xlsx'))
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    if not any(row): continue
    add(clean(row[0]), clean(row[1]), clean(row[2]), clean(row[3]))

# 5. TORTS_CIVIL.xlsx — cols: category, offense, ipc, bns, description
wb = openpyxl.load_workbook(os.path.join(folder, 'TORTS_CIVIL.xlsx'))
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    if not any(row): continue
    add(clean(row[0]), clean(row[1]), clean(row[2]), clean(row[3]), clean(row[4]))

# 6. WHITE_COLLAR.xlsx — cols: category, offense, ipc, bns
wb = openpyxl.load_workbook(os.path.join(folder, 'WHITE_COLLAR.xlsx'))
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    if not any(row): continue
    add(clean(row[0]), clean(row[1]), clean(row[2]), clean(row[3]))

with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

# Print summary
from collections import Counter
cats = Counter(d['category'] for d in data)
print(f'Total unique records: {len(data)}')
for cat, count in sorted(cats.items()):
    print(f'  {cat}: {count}')
